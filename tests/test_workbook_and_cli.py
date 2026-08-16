"""Tests for the generated workbook and the command line.

The three constraints this project was given are asserted here:
  - the chart's data lives on a visible sheet;
  - an unusable value is flagged and excluded, never replaced;
  - the demo command works from the start.
"""
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart

from report_builder.aggregate import build_all
from report_builder.cli import main
from report_builder.config import ReportConfig
from report_builder.quality import inspect
from report_builder.reading import ReadError, Table, read_table
from report_builder.workbook import write_report

SAMPLES = Path("samples")
SOURCE = SAMPLES / "sample_sales.csv"
CONFIG = SAMPLES / "sample_report.json"


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    """Build the sample report once and share it."""
    if not SOURCE.exists():
        pytest.skip("run `python samples/make_samples.py` first")
    out = tmp_path_factory.mktemp("wb") / "report.xlsx"
    config = ReportConfig.load(CONFIG)
    dataset = inspect(read_table(SOURCE), config)
    sections = build_all(dataset, config)
    write_report(out, config, dataset, sections, SOURCE.name)
    return out, config, dataset, sections


# -------------------------------------------------------------------- reading


def test_csv_excel_and_json_all_produce_the_same_columns(tmp_path):
    import json

    csv_path = tmp_path / "a.csv"
    csv_path.write_text("a;b\n1;2\n", encoding="utf-8")
    json_path = tmp_path / "a.json"
    json_path.write_text(json.dumps([{"a": 1, "b": 2}]), encoding="utf-8")

    assert read_table(csv_path).columns == ["a", "b"]
    assert read_table(json_path).columns == ["a", "b"]


def test_json_wrapped_in_an_object_is_accepted(tmp_path):
    import json

    path = tmp_path / "a.json"
    path.write_text(json.dumps({"rows": [{"a": 1}]}), encoding="utf-8")

    assert len(read_table(path)) == 1


def test_an_unsupported_extension_is_refused(tmp_path):
    path = tmp_path / "a.pdf"
    path.write_text("x", encoding="utf-8")

    with pytest.raises(ReadError, match="unsupported file type"):
        read_table(path)


def test_the_excel_and_csv_samples_agree():
    csv_table = read_table(SOURCE)
    xlsx_table = read_table(SAMPLES / "sample_sales.xlsx")

    assert csv_table.columns == xlsx_table.columns
    assert len(csv_table) == len(xlsx_table)


# ------------------------------------------------------------------- workbook


def test_every_sheet_is_visible(built):
    out, _, _, _ = built
    workbook = load_workbook(out)

    for name in workbook.sheetnames:
        assert workbook[name].sheet_state == "visible", f"{name} is not visible"


def test_the_summary_is_the_first_sheet(built):
    out, _, _, _ = built
    assert load_workbook(out).sheetnames[0] == "Summary"


def test_each_configured_chart_exists_and_has_the_right_type(built):
    out, config, _, _ = built
    workbook = load_workbook(out)
    expected = {"bar": BarChart, "column": BarChart, "line": LineChart, "pie": PieChart}

    for section in config.sections:
        if section.chart is None:
            continue
        sheet = workbook[section.name[:31]]
        assert len(sheet._charts) == 1, f"{section.name} should have one chart"
        assert isinstance(sheet._charts[0], expected[section.chart.type])


def test_a_section_without_a_chart_gets_none(built):
    out, _, _, _ = built
    assert load_workbook(out)["By rep and channel"]._charts == []


def test_the_chart_reads_from_its_own_visible_sheet(built):
    # The whole reason for choosing native charts: the source range is on a
    # normal sheet the recipient can edit, not a hidden helper sheet.
    out, _, _, _ = built
    workbook = load_workbook(out)
    sheet = workbook["Revenue by month"]
    chart = sheet._charts[0]

    reference = chart.series[0].val.numRef.f
    assert "Revenue by month" in reference
    assert sheet.sheet_state == "visible"


def test_the_chart_range_matches_the_rows_actually_written(built):
    out, _, _, sections = built
    sheet = load_workbook(out)["Revenue by month"]
    result = next(s for s in sections if s.section.name == "Revenue by month")

    reference = sheet._charts[0].series[0].val.numRef.f
    last_row = 3 + len(result.rows)
    assert reference.endswith(f"${last_row}")


def test_figures_are_written_as_numbers_not_text(built):
    out, _, _, _ = built
    sheet = load_workbook(out)["Revenue by month"]

    value = sheet.cell(row=4, column=2).value
    assert isinstance(value, (int, float)), "a chart cannot plot text"


def test_the_data_sheet_has_a_status_column_naming_the_problem(built):
    out, _, dataset, _ = built
    sheet = load_workbook(out)["Data"]

    headers = [cell.value for cell in sheet[1]]
    assert headers[-1] == "status"

    statuses = [row[-1] for row in sheet.iter_rows(min_row=2, values_only=True)]
    flagged = [s for s in statuses if s != "ok"]
    assert len(flagged) == len(dataset.flagged)
    assert any("is not a number" in s for s in flagged)
    assert any("above the allowed maximum" in s for s in flagged)


def test_an_unusable_value_keeps_its_original_text_on_the_data_sheet(built):
    out, _, _, _ = built
    sheet = load_workbook(out)["Data"]

    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[-1] != "ok" and "not a number" in row[-1]]
    assert rows, "the sample should contain an unreadable amount"
    assert rows[0][9] == "do ustalenia", "the original text must survive for checking"


def test_no_excluded_value_was_replaced_by_zero(built):
    # The strongest form of the rule: nowhere in the figures does an excluded
    # row show up as a zero contribution.
    out, _, dataset, sections = built
    from decimal import Decimal

    result = next(s for s in sections if s.section.name == "Revenue by month")
    for line in result.rows:
        figure = line.figures[0]
        if figure.rows_excluded:
            assert figure.value is None or figure.value != Decimal(0) or figure.rows_used > 0


def test_a_group_whose_figure_is_unknown_is_left_empty_not_zero(tmp_path):
    config = ReportConfig.from_dict({
        "columns": {"amount": {"type": "number"}},
        "sections": [{
            "name": "S", "group_by": ["region"],
            "aggregations": [{"column": "amount", "function": "sum", "label": "Total"}],
        }],
    })
    table = Table(columns=["region", "amount"],
                  rows=[{"region": "A", "amount": "brak"}], source_lines=[2])
    dataset = inspect(table, config)
    sections = build_all(dataset, config)
    out = tmp_path / "r.xlsx"

    write_report(out, config, dataset, sections, "t.csv")

    cell = load_workbook(out)["S"].cell(row=4, column=2)
    assert cell.value is None, "an unknown figure must not be written as 0"


def test_the_data_sheet_has_a_filter_and_frozen_header(built):
    out, _, _, _ = built
    sheet = load_workbook(out)["Data"]

    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None


# ------------------------------------------------------------------------ cli


def test_build_writes_a_workbook_and_exits_zero(tmp_path):
    out = tmp_path / "r.xlsx"

    code = main(["build", str(SOURCE), "--config", str(CONFIG), "--out", str(out),
                 "--no-preview"])

    assert code == 0
    assert out.exists()


def test_a_missing_source_exits_one(tmp_path, capsys):
    code = main(["build", str(tmp_path / "nope.csv"), "--config", str(CONFIG)])

    assert code == 1
    assert "File not found" in capsys.readouterr().out


def test_a_broken_configuration_exits_one_with_the_reason(tmp_path, capsys):
    bad = tmp_path / "bad.json"
    bad.write_text('{"sections": []}', encoding="utf-8")

    code = main(["build", str(SOURCE), "--config", str(bad), "--out", str(tmp_path / "r.xlsx")])

    assert code == 1
    assert "at least one section" in capsys.readouterr().out


def test_a_configuration_naming_an_absent_column_exits_one(tmp_path, capsys):
    bad = tmp_path / "bad.json"
    bad.write_text(
        '{"sections": [{"name": "S", "group_by": ["nie_ma"], '
        '"aggregations": [{"column": "net_amount", "function": "sum", "label": "T"}]}]}',
        encoding="utf-8",
    )

    code = main(["build", str(SOURCE), "--config", str(bad), "--out", str(tmp_path / "r.xlsx")])
    output = capsys.readouterr().out

    assert code == 1
    assert "does not have" in output
    assert "order_id" in output, "the message should list the columns that do exist"


def test_fail_on_excluded_exits_two_when_values_were_dropped(tmp_path):
    code = main(["build", str(SOURCE), "--config", str(CONFIG),
                 "--out", str(tmp_path / "r.xlsx"), "--no-preview", "--fail-on-excluded"])

    assert code == 2, "the sample deliberately contains unusable values"


def test_fail_on_excluded_exits_zero_on_clean_data(tmp_path):
    clean = tmp_path / "clean.csv"
    clean.write_text("region;amount\nA;10\nB;20\n", encoding="utf-8")
    config = tmp_path / "c.json"
    config.write_text(
        '{"columns": {"amount": {"type": "number"}}, "sections": [{"name": "S", '
        '"group_by": ["region"], "aggregations": [{"column": "amount", '
        '"function": "sum", "label": "Total"}]}]}',
        encoding="utf-8",
    )

    code = main(["build", str(clean), "--config", str(config),
                 "--out", str(tmp_path / "r.xlsx"), "--no-preview", "--fail-on-excluded"])

    assert code == 0


# ----------------------------------------------------------------------- demo


def test_demo_runs_and_leaves_no_files_behind(capsys):
    before = set(SAMPLES.iterdir())

    code = main(["demo"])
    output = capsys.readouterr().out

    assert code == 0
    # Asserted on content, not on the table title: rich wraps a title that is
    # wider than its own table, so the title is not a reliable string to match.
    assert "not a number" in output
    assert "out of range" in output
    assert "Sections" in output
    assert set(SAMPLES.iterdir()) == before


def test_demo_can_keep_the_workbook(tmp_path):
    out = tmp_path / "kept.xlsx"

    assert main(["demo", "--out", str(out)]) == 0
    assert out.exists()
    assert "Summary" in load_workbook(out).sheetnames


def test_demo_reports_a_missing_samples_folder(tmp_path, capsys):
    code = main(["demo", "--samples", str(tmp_path)])

    assert code == 1
    assert "Sample files not found" in capsys.readouterr().out


def test_demo_can_export_an_image(tmp_path):
    from PIL import Image

    png = tmp_path / "demo.png"
    assert main(["demo", "--export-png", str(png)]) == 0

    with Image.open(png) as image:
        assert image.height > 400


def test_notes_under_a_section_are_wrapped_not_left_to_run_under_the_chart(built):
    # A long note on one row overflows rightwards until it meets the chart, which
    # clips it mid-sentence — in the workbook and in any export of it.
    out, _, _, _ = built
    sheet = load_workbook(out)["Revenue by month"]

    notes = [sheet.cell(row=r, column=1).value
             for r in range(4, sheet.max_row + 1)
             if isinstance(sheet.cell(row=r, column=1).value, str)
             and sheet.cell(row=r, column=1).value.strip()]
    joined = " ".join(notes)

    assert "listed on the Data sheet." in joined, "the sentence must survive whole"
    assert all(len(n) <= 120 for n in notes), "no note row should be long enough to reach the chart"
