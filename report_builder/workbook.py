"""
Writing the workbook.

Layout, and why:

  - **Summary** — what the report is, what it was built from, and how much data
    each figure rests on. First sheet, so it is what opens.
  - **One sheet per section** — the aggregated table, with the chart anchored
    beside it. This table is the chart's data source and it is a normal, visible
    sheet: correcting a number here moves the chart.
  - **Data** — every source row, plus a `status` column saying whether the row
    was usable and, when it was not, exactly what was wrong with it.

No sheet is hidden and no range is protected. The point of shipping a workbook
rather than a picture is that the recipient can open it, check the numbers
against their own records and change them.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .aggregate import SectionResult
from .charts import build_chart
from .config import ReportConfig
from .quality import Dataset
from .values import as_datetime

_HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_FLAG_FILL = PatternFill("solid", fgColor="FFF3CD")
_MAX_WIDTH = 52


def write_report(path: Path, config: ReportConfig, dataset: Dataset,
                 sections: list[SectionResult], source_name: str) -> None:
    """Write the whole workbook to `path`."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Summary")
    used_names: set[str] = {"Summary"}

    for result in sections:
        name = _sheet_name(result.section.name, used_names)
        used_names.add(name)
        sheet = workbook.create_sheet(name)
        _write_section(sheet, result, config)

    data_sheet = workbook.create_sheet("Data")
    _write_data(data_sheet, dataset, config)

    _write_summary(summary, config, dataset, sections, source_name, used_names)
    workbook.save(path)


# --------------------------------------------------------------------- summary


def _write_summary(sheet, config: ReportConfig, dataset: Dataset,
                   sections: list[SectionResult], source_name: str,
                   sheet_names: set[str]) -> None:
    sheet["A1"] = config.title
    sheet["A1"].font = _TITLE_FONT

    facts = [
        ("Source file", source_name),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Rows read", dataset.total),
        ("Rows with something flagged", len(dataset.flagged)),
    ]
    row = 3
    for label, value in facts:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    tally = dataset.counts_by_verdict()
    if tally:
        row += 1
        sheet.cell(row=row, column=1, value="Values that could not be used").font = Font(bold=True)
        row += 1
        for reason, count in sorted(tally.items()):
            sheet.cell(row=row, column=1, value=reason)
            sheet.cell(row=row, column=2, value=count)
            row += 1
        sheet.cell(row=row, column=1,
                   value="These values were excluded from the figures below. "
                         "None of them was replaced or estimated.")
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Sections").font = Font(bold=True)
    row += 1
    headers = ["Section", "Groups", "Values behind the figures", "Values excluded", "Notes"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    row += 1

    for result in sections:
        used = sum(f.rows_used for line in result.rows for f in line.figures)
        notes = []
        if result.rows_without_key:
            notes.append(
                f"{result.rows_without_key} row(s) had no usable grouping value"
            )
        notes.extend(result.key_notes)
        sheet.cell(row=row, column=1, value=result.section.name)
        sheet.cell(row=row, column=2, value=len(result.rows))
        sheet.cell(row=row, column=3, value=used)
        excluded = result.excluded_total + result.rows_without_key
        cell = sheet.cell(row=row, column=4, value=excluded)
        if excluded:
            cell.fill = _FLAG_FILL
        sheet.cell(row=row, column=5, value="; ".join(notes))
        row += 1

    row += 1
    sheet.cell(row=row, column=1,
               value="Each section sheet holds the table its chart is drawn from. "
                     "Edit a figure there and the chart follows.")
    _fit(sheet)


# -------------------------------------------------------------------- sections


def _write_section(sheet, result: SectionResult, config: ReportConfig) -> None:
    sheet["A1"] = result.section.name
    sheet["A1"].font = _TITLE_FONT

    header_row = 3
    headers = result.key_labels + result.figure_labels()
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    key_count = len(result.key_labels)
    row_number = header_row
    for line in result.rows:
        row_number += 1
        for offset, key in enumerate(line.keys, start=1):
            sheet.cell(row=row_number, column=offset, value=key)
        for offset, figure in enumerate(line.figures, start=key_count + 1):
            cell = sheet.cell(row=row_number, column=offset)
            if figure.value is None:
                # No usable value at all. Left visibly empty with a note rather
                # than written as 0, which would read as a real measurement.
                cell.value = None
                cell.comment = _comment("no usable value in this group")
            elif isinstance(figure.value, Decimal):
                cell.value = float(figure.value)
                cell.number_format = config.number_format
            else:
                cell.value = figure.value
            if figure.rows_excluded:
                cell.fill = _FLAG_FILL
                cell.comment = _comment(
                    f"{figure.rows_excluded} row(s) excluded — see the Data sheet"
                )

    last_row = row_number
    value_columns = {
        label: key_count + 1 + index
        for index, label in enumerate(result.figure_labels())
    }

    if result.section.chart and result.rows:
        chart = build_chart(
            result.section.chart, sheet, header_row, last_row, key_count, value_columns
        )
        if chart is not None:
            anchor = f"{get_column_letter(len(headers) + 2)}{header_row}"
            sheet.add_chart(chart, anchor)

    # Column widths must be settled before the notes are written, because the
    # notes are wrapped to fit inside them.
    _fit(sheet)

    notes = ["This table is the chart's data source. Change a value and the chart updates."]
    if result.rows_without_key:
        notes.append(
            f"{result.rows_without_key} row(s) are not represented here: their grouping "
            "value could not be read. They are listed on the Data sheet."
        )
    _write_notes(sheet, notes, last_row + 2, len(headers))


# ------------------------------------------------------------------------ data


def _write_data(sheet, dataset: Dataset, config: ReportConfig) -> None:
    columns = list(dataset.columns)
    headers = columns + ["status"]
    for index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT

    for offset, row in enumerate(dataset.rows, start=2):
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(row=offset, column=index)
            _write_value(cell, row, column, config)
        status = row.status
        status_cell = sheet.cell(row=offset, column=len(columns) + 1, value=status)
        if status != "ok":
            status_cell.fill = _FLAG_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(dataset.rows) + 1)}"
    _fit(sheet)


def _write_value(cell, row, column: str, config: ReportConfig) -> None:
    """Write one source value, typed when it was usable, raw text when it was not.

    An unusable value keeps its original text so the recipient can see exactly
    what the file contained — that is what makes the status column checkable.
    """
    parsed = row.cells.get(column)
    if parsed is None:
        cell.value = row.raw.get(column, "")
        return
    if not parsed.usable:
        cell.value = parsed.raw
        cell.fill = _FLAG_FILL
        return
    value = parsed.value
    if isinstance(value, Decimal):
        cell.value = float(value)
        cell.number_format = config.number_format
    elif isinstance(value, date):
        cell.value = as_datetime(value)
        cell.number_format = config.date_format
    else:
        cell.value = value


# ---------------------------------------------------------------------- shared


def _write_notes(sheet, notes: list[str], first_row: int, table_columns: int) -> None:
    """Write the notes under the table, wrapped to the table's own width.

    A long note left on one row overflows to the right until it meets something,
    and what it meets is the chart — which clips it mid-sentence, in the workbook
    as well as in any print or PDF export of it. Wrapping to the width the table
    actually occupies keeps every sentence readable.
    """
    import textwrap

    # The columns the note may spill across: the table, plus the empty gap column
    # before the chart's anchor.
    budget = 0
    for column in range(1, table_columns + 2):
        dimension = sheet.column_dimensions.get(get_column_letter(column))
        budget += int(dimension.width) if dimension and dimension.width else 10
    # Excel column "width" is measured in characters of the default font, so the
    # sum is already a character budget. A small margin avoids touching the chart.
    budget = max(40, budget - 4)

    row = first_row
    for note in notes:
        for line in textwrap.wrap(note, width=budget) or [""]:
            sheet.cell(row=row, column=1, value=line)
            row += 1


def _comment(text: str):
    from openpyxl.comments import Comment

    return Comment(text, "report-builder")


def _sheet_name(name: str, taken: set[str]) -> str:
    """Excel sheet names are at most 31 characters and cannot repeat."""
    for character in "[]:*?/\\":
        name = name.replace(character, "-")
    candidate = name[:31] or "Section"
    suffix = 2
    while candidate in taken:
        trimmed = name[:28]
        candidate = f"{trimmed}-{suffix}"
        suffix += 1
    return candidate


def _fit(sheet) -> None:
    """Size the columns so the recipient does not have to widen them by hand."""
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if isinstance(cell.value, datetime):
                length = 10
            widths[cell.column] = min(max(widths.get(cell.column, 0), length + 2), _MAX_WIDTH)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
