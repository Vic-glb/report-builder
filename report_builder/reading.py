"""
Reading the source data.

CSV, Excel and JSON go in, a table of raw strings comes out. As in the sibling
tools, nothing is converted here: the original text is what the status column
later quotes back to the reader, so it has to survive this far intact.
"""
from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from pathlib import Path

#: Tried in order. cp1250 and iso-8859-2 cover Central European exports.
ENCODINGS = ("utf-8-sig", "utf-8", "cp1250", "iso-8859-2", "cp1252", "latin-1")
DELIMITERS = (";", ",", "\t", "|")


@dataclass
class Table:
    """Source rows as raw strings, keyed by column name."""

    columns: list[str]
    rows: list[dict[str, str]] = field(default_factory=list)
    #: 1-based line number in the source, so a status message can point at it.
    source_lines: list[int] = field(default_factory=list)
    #: Notes about how the file was read (encoding, delimiter).
    notes: list[str] = field(default_factory=list)

    def __len__(self) -> int:
        return len(self.rows)


class ReadError(ValueError):
    """Raised when the source file cannot be turned into a table."""


def read_table(path: Path, sheet: str | None = None) -> Table:
    """Read a `.csv`, `.xlsx` or `.json` file.

    Raises:
        ReadError: If the extension is unsupported, the file is empty, or a JSON
            document is not a list of objects.
    """
    suffix = path.suffix.lower()
    try:
        if suffix in (".csv", ".txt", ".tsv"):
            return _read_csv(path)
        if suffix in (".xlsx", ".xlsm"):
            return _read_excel(path, sheet)
        if suffix == ".json":
            return _read_json(path)
    except OSError as exc:
        raise ReadError(f"{path.name} could not be read: {exc}") from exc
    raise ReadError(f"{path.name}: unsupported file type {suffix or '(none)'}")


def _read_csv(path: Path) -> Table:
    data = path.read_bytes()
    text, encoding = _decode(data)
    delimiter = _sniff(text)

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    grid = [row for row in reader]
    if not grid:
        raise ReadError(f"{path.name} is empty")

    header = [cell.strip() for cell in grid[0]]
    header = _unique(header)
    rows, lines = [], []
    for number, row in enumerate(grid[1:], start=2):
        padded = list(row) + [""] * (len(header) - len(row))
        record = dict(zip(header, padded))
        if not any(value.strip() for value in record.values()):
            continue
        rows.append(record)
        lines.append(number)

    notes = [f"Delimiter detected as {_delimiter_name(delimiter)}."]
    if encoding not in ("utf-8", "utf-8-sig"):
        notes.append(f"File decoded as {encoding}, not UTF-8.")
    return Table(columns=header, rows=rows, source_lines=lines, notes=notes)


def _read_excel(path: Path, sheet: str | None) -> Table:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active
    except KeyError as exc:
        raise ReadError(f"{path.name} has no sheet named {sheet!r}") from exc

    grid = [
        ["" if cell is None else _cell_text(cell) for cell in row]
        for row in worksheet.iter_rows(values_only=True)
    ]
    workbook.close()
    if not grid:
        raise ReadError(f"{path.name} is empty")

    header = _unique([cell.strip() for cell in grid[0]])
    rows, lines = [], []
    for number, row in enumerate(grid[1:], start=2):
        padded = list(row) + [""] * (len(header) - len(row))
        record = dict(zip(header, padded))
        if not any(value.strip() for value in record.values()):
            continue
        rows.append(record)
        lines.append(number)
    return Table(columns=header, rows=rows, source_lines=lines,
                 notes=[f"Read from worksheet {worksheet.title!r}."])


def _read_json(path: Path) -> Table:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ReadError(f"{path.name} is not valid JSON: {exc}") from exc

    if isinstance(payload, dict):
        # A common shape: {"rows": [...]} or {"data": [...]}.
        for key in ("rows", "data", "items", "records"):
            if isinstance(payload.get(key), list):
                payload = payload[key]
                break
    if not isinstance(payload, list) or not payload:
        raise ReadError(
            f"{path.name}: expected a list of objects, or an object holding one "
            'under "rows", "data", "items" or "records"'
        )
    if not all(isinstance(entry, dict) for entry in payload):
        raise ReadError(f"{path.name}: every entry must be an object")

    columns: list[str] = []
    for entry in payload:
        for key in entry:
            if key not in columns:
                columns.append(key)

    rows, lines = [], []
    for number, entry in enumerate(payload, start=1):
        rows.append({column: _text(entry.get(column)) for column in columns})
        lines.append(number)
    return Table(columns=columns, rows=rows, source_lines=lines,
                 notes=[f"Read {len(rows)} records from JSON."])


def _decode(data: bytes) -> tuple[str, str]:
    for encoding in ENCODINGS:
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8 (with replacements)"


def _sniff(text: str) -> str:
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters="".join(DELIMITERS)).delimiter
    except csv.Error:
        pass
    best, best_score = ",", -1.0
    for candidate in DELIMITERS:
        counts = [len(row) for row in csv.reader(io.StringIO(sample), delimiter=candidate)]
        counts = [c for c in counts if c > 0]
        if not counts or max(counts) < 2:
            continue
        common = max(set(counts), key=counts.count)
        score = (counts.count(common) / len(counts)) * max(counts)
        if score > best_score:
            best, best_score = candidate, score
    return best


def _delimiter_name(delimiter: str) -> str:
    return {";": "semicolon", ",": "comma", "\t": "tab", "|": "pipe"}.get(delimiter, delimiter)


def _unique(header: list[str]) -> list[str]:
    """Give every column a distinct, non-empty name."""
    seen: dict[str, int] = {}
    names: list[str] = []
    for index, name in enumerate(header, start=1):
        candidate = name or f"column_{index}"
        base = candidate
        while candidate.lower() in seen:
            seen[base.lower()] = seen.get(base.lower(), 1) + 1
            candidate = f"{base}_{seen[base.lower()]}"
        seen[candidate.lower()] = 1
        names.append(candidate)
    return names


def _cell_text(cell) -> str:
    from datetime import date as _date, datetime as _datetime

    if isinstance(cell, _datetime):
        return cell.date().isoformat() if cell.time().isoformat() == "00:00:00" else cell.isoformat(" ")
    if isinstance(cell, _date):
        return cell.isoformat()
    if isinstance(cell, float) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)
